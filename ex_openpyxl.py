import os
import sys
import glob
import datetime
import openpyxl
import numpy as np


def main():
    # ----- open -----
    # create a new workbook
    wb = openpyxl.Workbook()
    # load a workbook
    """
    file_path = input("--> ")
    wb = openpyxl.load_workbook(file_path, data_only=True)
    """

    # ----- select a worksheet -----
    # the active sheet
    ws = wb.active
    # using the sheet name
    """
    ws_name = wb.sheetnames[0]
    ws = wb[ws_name]
    """

    # ----- write -----
    # write to a cell
    ws["A1"].value = 1
    ws.cell(row=1, column=2, value=2)
    # write a list
    list_1d = [3, 4, 5, 6, 7]
    ws.append(list_1d)
    list_2d = [[8, 9, 10, 11, 12], [13, 14, 15, 16, 17]]
    for i in list_2d: ws.append(i)

    # ----- cell style -----
    ws["A1"].fill = openpyxl.styles.PatternFill(patternType="solid", fgColor="ffaaaa")
    ws["A1"].number_format = "0.00"
    ws["B1"].font = openpyxl.styles.Font(size=20, bold=True, italic=True, color="0000ff")

    # ----- read -----
    # read from a cell
    val = ws["A1"].value
    print(val)
    val = ws.cell(row=1, column=2).value
    print(val)
    # read from in a range
    val = ws["A1:D4"]
    val = [[j.value for j in i] for i in val]
    for i in val: print(i)

    # ----- control sheets -----
    """
    ws.title = "test_sheet"
    ws_add = wb.create_sheet(title="new_sheet", index=0)
    ws_copy = wb.copy_worksheet(ws_add)
    wb.remove(ws_add)
    """

    # ----- graph -----
    # create a new sheet
    ws_add = wb.create_sheet(title="graph")
    ws = ws_add
    # write x,y datas
    n = 10 # number of datas = n-1
    x = np.arange(1,n)
    y1 = np.random.randint(1,10,(n-1))
    y2 = np.random.randint(10,20,(n-1))
    data = np.vstack((x, y1, y2))
    data = data.transpose()
    data = np.vstack((np.array(["x", "y1", "y2"]), data))
    for i in data.tolist(): ws.append(i)
    # draw a scatter graph
    chart = openpyxl.chart.ScatterChart()
    chart.title = "scatter graph"
    chart.x_axis.title = "x_axis"
    chart.y_axis.title = "y_axis"

    xval = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=10)
    yval = openpyxl.chart.Reference(ws, min_col=2, min_row=2, max_row=10)
    series = openpyxl.chart.Series(yval, xval, title="sample1")

    series.marker.symbol = "circle"
    series.marker.size = 10
    series.marker.spPr.fill = "0000ff"

    chart.series.append(series)

    """
    yval = openpyxl.chart.Reference(ws, min_col=3, min_row=2, max_row=10)
    series = openpyxl.chart.Series(y2val, xval)
    chart.series.append(series)
    """

    ws.add_chart(chart, "D1")

    # save
    wb.save("test.xlsx")

    # close
    wb.close()


def make_dir():
    # directory path(name)
    file_path = os.path.abspath(__file__)
    dir_path = os.path.dirname(file_path)
    file_name = os.path.basename(file_path)
    date = datetime.datetime.now().strftime("%Y%m%d")
    dir_name = date + "_" + file_name[:-3]
    dir_path = os.path.join(dir_path, dir_name)
    # create a directory
    if not os.path.exists(dir_path): os.mkdir(dir_path)
    # move to the directory
    os.chdir(dir_path)


if __name__ == "__main__":
    # run the main function
    print("\n", "----- start -----", "\n")
    make_dir()
    main()
    print("\n", "----- end -----", "\n")
