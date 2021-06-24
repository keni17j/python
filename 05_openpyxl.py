"""
Openpyxl.
"""

import os
import sys

import numpy as np
import openpyxl


def main():

    # Create a new workbook.
    wb = openpyxl.Workbook()
    # Load a workbook.
    #file_path = input("--> ")
    #wb = openpyxl.load_workbook(file_path, data_only=True)

    # Select the worksheet.
    ws = wb.active  # Active sheet.
    #ws_name = wb.sheetnames[0]
    #ws = wb[ws_name]  # Select by the sheet name.

    write(ws)
    read(ws)
    edit(wb, ws)
    graph(wb, ws)

    # Save.
    wb.save('test.xlsx')

    # Close.
    wb.close()


def write(ws):

    # Write a data.
    ws['A1'].value = 1
    ws.cell(row=1, column=2, value=2)

    # Write a list.
    list_1d = [3, 4, 5, 6, 7]
    ws.append(list_1d)
    list_2d = [[8, 9, 10, 11, 12], [13, 14, 15, 16, 17]]
    for i in list_2d:
        ws.append(i)


def read(ws):

    # Read from the cell.
    val = ws['A1'].value
    print(val)
    val = ws.cell(row=1, column=2).value
    print(val)
    val = ws['A1:D4']
    val = list(val)  # Convert tuples to lists because tuples can't be overwrited.
    for i, row in enumerate(val):
        val[i] = [j.value for j in row]  # Convert to values.
    for i in val: print(i)


def edit(wb, ws):

    # Cell style.
    # Fill.
    ws['A1'].fill = openpyxl.styles.PatternFill(patternType='solid', fgColor='ffaaaa')
    # Format.
    ws['A1'].number_format = '0.00'
    # Font.
    ws['B1'].font = openpyxl.styles.Font(size=20, bold=True, italic=True, color='0000ff')

    # Sheets.
    # Title.
    ws.title = 'test_sheet'
    # Add a new sheet.
    ws_add = wb.create_sheet(title='new_sheet', index=0)
    # Copy the sheet.
    ws_copy = wb.copy_worksheet(ws_add)
    # Delete the sheet.
    wb.remove(ws_add)
    wb.remove(ws_copy)


def graph(wb, ws):

    # Create a new sheet.
    ws = wb.create_sheet(title='graph')
    # Write x,y datas.
    n = 10  # Number of datas, n - 1
    x = np.arange(1, n)
    y1 = np.random.randint(1, 10, n-1)
    y2 = np.random.randint(10, 20, n-1)
    data = np.vstack((x, y1, y2))
    data = data.transpose()
    data = np.vstack((['x', 'y1', 'y2'], data))
    for i in data.tolist():
        ws.append(i)

    # Draw a scatter graph.
    chart = openpyxl.chart.ScatterChart()
    # Titles.
    chart.title = 'scatter graph'
    chart.x_axis.title = 'x_axis'
    chart.y_axis.title = 'y_axis'
    # Select values.
    xval = openpyxl.chart.Reference(ws, min_col=1, min_row=2, max_row=10)
    yval = openpyxl.chart.Reference(ws, min_col=2, min_row=2, max_row=10)
    series = openpyxl.chart.Series(yval, xval, title='sample1')
    # Setting.
    series.marker.symbol = 'circle'
    series.marker.size = 10
    series.marker.spPr.fill = '0000ff'
    chart.series.append(series)

    # Add datas.
    yval = openpyxl.chart.Reference(ws, min_col=3, min_row=2, max_row=10)
    series = openpyxl.chart.Series(yval, xval)
    chart.series.append(series)

    # create the graph
    ws.add_chart(chart, 'D1')


if __name__ == '__main__':
    main()
